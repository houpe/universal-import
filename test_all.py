#!/usr/bin/env python3
"""
万能导入系统 - 全量测试用例
基于考试系统.md的所有考点要求
"""

import requests
import json
import time
import sys
from io import BytesIO

BASE_URL = "https://exam-project-fawn.vercel.app"

# ==================== 颜色输出 ====================
class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    END = '\033[0m'
    BOLD = '\033[1m'

def log_pass(msg):
    print(f"  {Colors.GREEN}✅ PASS{Colors.END} {msg}")

def log_fail(msg):
    print(f"  {Colors.RED}❌ FAIL{Colors.END} {msg}")

def log_warn(msg):
    print(f"  {Colors.YELLOW}⚠️  WARN{Colors.END} {msg}")

def log_info(msg):
    print(f"  {Colors.BLUE}ℹ️  INFO{Colors.END} {msg}")

def log_section(title):
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE} {title}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")

# ==================== 考点1: Vercel部署 ====================
def test_tc_1_deployment():
    log_section("考点1: Vercel部署 (前置条件)")

    # TC-1.1: Vercel可访问
    try:
        r = requests.get(BASE_URL, timeout=15)
        if r.status_code == 200:
            log_pass(f"TC-1.1: Vercel可访问 (status={r.status_code})")
        else:
            log_fail(f"TC-1.1: Vercel返回status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-1.1: Vercel无法访问 - {e}")

    # TC-1.2: 技术栈 - Next.js App Router
    try:
        r = requests.get(f"{BASE_URL}/api/orders?page=1&pageSize=1", timeout=10)
        if r.status_code == 200:
            data = r.json()
            if "data" in data and "total" in data:
                log_pass("TC-1.2: API路由正常工作，Next.js App Router已部署")
            else:
                log_fail("TC-1.2: API响应格式异常")
        else:
            log_fail(f"TC-1.2: API返回status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-1.2: API无法访问 - {e}")

# ==================== 考点2: Excel解析与多模板支持 ====================
def test_tc_2_excel_parsing():
    log_section("考点2: Excel解析与多模板支持 (46分)")

    templates = [
        ("template1-standard.xlsx", "标准模板"),
        ("template2-ecommerce.xlsx", "电商模板（合并单元格+说明行）"),
        ("template3-english.xlsx", "英文列名模板"),
        ("template4-grouped.xlsx", "分组表头模板（2行表头）"),
        ("template5-multisheet.xlsx", "多Sheet模板"),
    ]

    for filename, desc in templates:
        try:
            with open(f"public/templates/{filename}", "rb") as f:
                files = {"file": (filename, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
                r = requests.post(f"{BASE_URL}/api/parse", files=files, timeout=30)

            if r.status_code == 200:
                data = r.json()
                headers = data.get("headers", [])
                rows = data.get("rows", [])
                auto_mapping = data.get("autoMapping", {})
                mapped_data = data.get("mappedData", [])

                # 检查autoMapping中有多少个字段被成功映射
                mapped_count = sum(1 for k, v in auto_mapping.items() if v is not None)
                row_count = len(rows)

                if mapped_count >= 9:  # 至少9个字段映射成功
                    log_pass(f"TC-2.x {desc}: {mapped_count}/11 字段映射成功, {row_count} 行数据")
                    # 打印映射结果
                    log_info(f"  原始表头: {headers}")
                    log_info(f"  映射关系: {auto_mapping}")
                    if mapped_data:
                        log_info(f"  示例数据: {mapped_data[0]}")
                else:
                    log_fail(f"TC-2.x {desc}: 仅 {mapped_count}/11 字段映射成功")
                    log_info(f"  原始表头: {headers}")
                    log_info(f"  映射关系: {auto_mapping}")
            else:
                log_fail(f"TC-2.x {desc}: 解析API返回status={r.status_code}")
        except Exception as e:
            log_fail(f"TC-2.x {desc}: 异常 - {e}")

    # TC-2.10: 大文件处理 - 生成1000+行测试
    log_info("\nTC-2.10: 生成1000行测试数据...")
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "大文件测试"
        # 表头
        ws.append(["外部编码", "发件人姓名", "发件人电话", "发件人地址",
                    "收件人姓名", "收件人电话", "收件人地址", "重量(kg)", "件数", "温层", "备注"])
        # 1000行数据
        for i in range(1, 1001):
            ws.append([
                f"BIG-TEST-{i:04d}", f"发件人{i}", "13800138001", "北京市朝阳区",
                f"收件人{i}", "13900139001", "上海市浦东新区",
                str(1 + i % 10), str(1 + i % 5), "常温", f"测试备注{i}"
            ])
        big_file = "/tmp/big_test_1000.xlsx"
        wb.save(big_file)

        with open(big_file, "rb") as f:
            files = {"file": ("big_test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            start = time.time()
            r = requests.post(f"{BASE_URL}/api/parse", files=files, timeout=60)
            elapsed = time.time() - start

        if r.status_code == 200:
            data = r.json()
            row_count = len(data.get("rows", []))
            if row_count == 1000:
                log_pass(f"TC-2.10: 1000行解析成功, 耗时{elapsed:.2f}s")
            else:
                log_fail(f"TC-2.10: 仅解析 {row_count}/1000 行")
        else:
            log_fail(f"TC-2.10: 大文件解析失败, status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-2.10: 大文件测试异常 - {e}")

# ==================== 考点3: 错误处理与数据校验 ====================
def test_tc_3_validation():
    log_section("考点3: 错误处理与数据校验 (15分)")

    # TC-3.1~3.7: 通过预览页数据校验
    # 构造包含各种错误的测试数据
    test_rows = [
        # 行1: 正常数据
        ["ORD-001", "张三", "13800138001", "北京市朝阳区",
         "李四", "13900139001", "上海市浦东新区", "5.2", "2", "常温", "正常"],
        # 行2: 缺少必填字段(发件人姓名、收件人电话)
        ["ORD-002", "", "13800138002", "广州市天河区",
         "赵六", "", "深圳市南山区", "3", "1", "冷藏", ""],
        # 行3: 电话格式错误
        ["ORD-003", "王五", "12345", "北京市西城区",
         "孙七", "abc123", "天津市和平区", "2", "1", "冷冻", ""],
        # 行4: 重量/件数非正数
        ["ORD-004", "周八", "13800138004", "杭州市西湖区",
         "吴九", "13900139004", "南京市鼓楼区", "-5", "0", "常温", ""],
        # 行5: 温层不在范围内
        ["ORD-005", "郑十", "13800138005", "成都市武侯区",
         "王麻", "13900139005", "重庆市渝中区", "3", "2", "高温", ""],
        # 行6: 外部编码重复(与行1重复)
        ["ORD-001", "陈一", "13800138006", "武汉市江汉区",
         "刘二", "13900139006", "长沙市岳麓区", "4", "1", "常温", "重复编码"],
    ]

    # 通过API提交校验测试
    headers = ["外部编码", "发件人姓名", "发件人电话", "发件人地址",
               "收件人姓名", "收件人电话", "收件人地址", "重量(kg)",
               "件数", "温层", "备注"]

    try:
        # 构造Excel文件
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "校验测试"
        ws.append(headers)
        for row in test_rows:
            ws.append(row)
        test_file = "/tmp/validation_test.xlsx"
        wb.save(test_file)

        with open(test_file, "rb") as f:
            files = {"file": ("validation_test.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = requests.post(f"{BASE_URL}/api/parse", files=files, timeout=15)

        if r.status_code == 200:
            data = r.json()
            rows = data.get("rows", [])
            log_pass(f"TC-3.x: 解析成功，{len(rows)}行数据返回前端用于校验")
            log_info(f"  包含: 正常行、缺必填、电话错误、重量错误、温层错误、编码重复")
        else:
            log_fail(f"TC-3.x: 解析失败, status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-3.x: 校验测试异常 - {e}")

    # TC-3.9: 空文件
    log_info("\nTC-3.9: 测试空文件处理...")
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        empty_file = "/tmp/empty_test.xlsx"
        wb.save(empty_file)

        with open(empty_file, "rb") as f:
            files = {"file": ("empty.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = requests.post(f"{BASE_URL}/api/parse", files=files, timeout=10)

        if r.status_code == 200:
            data = r.json()
            if data.get("error"):
                log_pass(f"TC-3.9: 空文件友好提示: {data['error']}")
            else:
                log_warn(f"TC-3.9: 空文件返回空数据，未报错")
        elif r.status_code == 400:
            data = r.json()
            if data.get("error"):
                log_pass(f"TC-3.9: 空文件友好提示(400): {data['error']}")
            else:
                log_fail(f"TC-3.9: 空文件返回400但无错误信息")
        else:
            log_fail(f"TC-3.9: 空文件处理异常, status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-3.9: 空文件测试异常 - {e}")

    # TC-3.10: 格式错误文件
    log_info("\nTC-3.10: 测试非Excel文件...")
    try:
        fake_file = "/tmp/fake_test.txt"
        with open(fake_file, "w") as f:
            f.write("This is not an Excel file")

        with open(fake_file, "rb") as f:
            files = {"file": ("fake.txt", f, "text/plain")}
            r = requests.post(f"{BASE_URL}/api/parse", files=files, timeout=10)

        if r.status_code == 400 or (r.status_code == 200 and r.json().get("error")):
            log_pass(f"TC-3.10: 非Excel文件格式错误提示")
        else:
            log_warn(f"TC-3.10: 非Excel文件处理返回status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-3.10: 格式错误测试异常 - {e}")

# ==================== 考点6: 数据库存储 ====================
def test_tc_6_database():
    log_section("考点6: 数据库存储与运单查看 (10分)")

    # TC-6.1: 提交数据到数据库
    log_info("TC-6.1: 测试数据库写入...")
    try:
        test_orders = [
            {
                "external_code": "DB-TEST-001",
                "sender_name": "数据库测试发件人",
                "sender_phone": "13800138001",
                "sender_address": "北京市朝阳区",
                "receiver_name": "数据库测试收件人",
                "receiver_phone": "13900139001",
                "receiver_address": "上海市浦东新区",
                "weight": "5.2",
                "quantity": "2",
                "temp_zone": "常温",
                "remark": "数据库写入测试"
            }
        ]

        r = requests.post(
            f"{BASE_URL}/api/orders",
            json={"orders": test_orders},
            headers={"Content-Type": "application/json"},
            timeout=15
        )

        if r.status_code == 200:
            data = r.json()
            if data.get("success", 0) > 0:
                log_pass(f"TC-6.1: 数据库写入成功, success={data['success']}")
            else:
                log_fail(f"TC-6.1: 数据库写入失败")
        else:
            log_fail(f"TC-6.1: 写入API返回status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-6.1: 数据库测试异常 - {e}")

    # TC-6.2: 查询数据库
    log_info("\nTC-6.2: 测试数据库读取...")
    try:
        r = requests.get(f"{BASE_URL}/api/orders?page=1&pageSize=10", timeout=10)
        if r.status_code == 200:
            data = r.json()
            total = data.get("total", 0)
            orders = data.get("data", [])
            log_pass(f"TC-6.2: 数据库读取成功, 总计{total}条, 当前页{len(orders)}条")
        else:
            log_fail(f"TC-6.2: 读取API返回status={r.status_code}")
    except Exception as e:
        log_fail(f"TC-6.2: 数据库读取异常 - {e}")

    # TC-6.3~6.5: 筛选功能
    log_info("\nTC-6.3~6.5: 测试筛选功能...")
    try:
        # 按外部编码筛选
        r = requests.get(f"{BASE_URL}/api/orders?page=1&pageSize=10&external_code=DB-TEST", timeout=10)
        if r.status_code == 200:
            data = r.json()
            log_pass(f"TC-6.3: 按外部编码筛选, 找到{data.get('total', 0)}条")

        # 按收件人筛选
        r = requests.get(f"{BASE_URL}/api/orders?page=1&pageSize=10&receiver=测试收件人", timeout=10)
        if r.status_code == 200:
            data = r.json()
            log_pass(f"TC-6.4: 按收件人筛选, 找到{data.get('total', 0)}条")
    except Exception as e:
        log_fail(f"TC-6.3~6.5: 筛选测试异常 - {e}")

# ==================== 主测试流程 ====================
if __name__ == "__main__":
    print(f"\n{Colors.BOLD}万能导入系统 - 全量自动化测试{Colors.END}")
    print(f"目标地址: {BASE_URL}")
    print(f"开始时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")

    test_tc_1_deployment()
    test_tc_2_excel_parsing()
    test_tc_3_validation()
    test_tc_6_database()

    print(f"\n{Colors.BOLD}{Colors.GREEN}{'='*60}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.GREEN} 测试完成!{Colors.END}")
    print(f"{Colors.BOLD}{Colors.GREEN}{'='*60}{Colors.END}")
    print(f"结束时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"\n请同时通过浏览器验证:")
    print(f"  1. {BASE_URL}/import - 文件上传+解析")
    print(f"  2. {BASE_URL}/preview - 预览+编辑+校验")
    print(f"  3. {BASE_URL}/orders - 运单列表+筛选+分页")
